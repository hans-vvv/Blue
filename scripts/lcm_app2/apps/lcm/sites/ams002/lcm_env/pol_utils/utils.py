import re, os, pathlib, json, fnmatch, shutil
from openpyxl.utils import get_column_letter
import openpyxl


class ReSearcher():
    """
    Helper  to enable evaluation
    and regex formatting in a single line
    """

    match = None

    def __call__(self, pattern, string):
        self.match = re.search(pattern, string)
        return self.match

    def __getattr__(self, name):
        return getattr(self.match, name)


class Tree(dict):
    """ Autovivificious dictionary """

    def __missing__(self, key):
        value = self[key] = type(self)()
        return value

    def __str__(self):
        """ Serialize dictionary to JSON formatted string with indents """
        return json.dumps(self, indent=4)


def xlref(row, column, zero_indexed=True):
    """
    openpyxl helper
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def get_tree_by_hostname(trees, hostname):
    for tree in trees:
        if tree['hostname'] == hostname:
            return tree


def move_dir(src, dst, pattern='*'):
    """
    Function moves files from source to destination folder based on pattern
    of filename
    """
    if not os.path.isdir(dst):
        pathlib.Path(dst).mkdir(parents=True, exist_ok=True)
    for file in fnmatch.filter(os.listdir(src), pattern):
        shutil.move(os.path.join(src, file), os.path.join(dst, file))


def xls_writer(trees, filename):
    """
    Function prints trees object to excel file. Vlan
    and port info is printed in different tabs.
    """

    # Calculate set of keys.
    vlankeys = set()
    portkeys = set()

    for tree in trees:
        for vlan, vlanitems in tree['vlan'].items():
            vlankeys.update(vlanitems.keys())
        for port, portitems in tree['port'].items():
            portkeys.update(portitems.keys())

    vlankeys = sorted(set(vlankeys))
    portkeys = sorted(set(portkeys))
    if 'vlanindex' in vlankeys:
        vlankeys.remove('vlanindex')
    if 'name' in vlankeys:
        vlankeys.remove('name')
    if 'description' in portkeys:
        portkeys.remove('description')
    if 'portindex' in portkeys:
        portkeys.remove('portindex')
    if 'switchport mode' in portkeys:
        portkeys.remove('switchport mode')
    if 'switchport access vlan' in portkeys:
        portkeys.remove('switchport access vlan')
    if 'switchport voice vlan' in portkeys:
        portkeys.remove('switchport voice vlan')
    if 'vlan_allow_list' in portkeys:
        portkeys.remove('vlan_allow_list')
    if 'vrf forwarding' in portkeys:
        portkeys.remove('vrf forwarding')
    if 'speed' in portkeys:
        portkeys.remove('speed')
    if 'duplex' in portkeys:
        portkeys.remove('duplex')
    vlankeys.insert(0, 'name')
    portkeys.insert(0, 'vlan_allow_list')
    portkeys.insert(0, 'duplex')
    portkeys.insert(0, 'speed')
    portkeys.insert(0, 'vrf forwarding')
    portkeys.insert(0, 'switchport voice vlan')
    portkeys.insert(0, 'switchport access vlan')
    portkeys.insert(0, 'switchport mode')
    portkeys.insert(0, 'description')

    wb = openpyxl.load_workbook(filename)
    sheetnames = wb.sheetnames
    if 'OldVlaninfo' in sheetnames:
        wb.remove(wb['OldVlaninfo'])
    wb.create_sheet('OldVlaninfo')
    ws = wb['OldVlaninfo']

    count_vlan_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'vlanindex'
    for count, vlankey in enumerate(vlankeys):
        ws[xlref(0, count + 2)] = vlankey

    for tree in trees:
        for vlan, vlanitems in tree['vlan'].items():
            ws[xlref(count_vlan_row + 1, 0)] = tree['hostname']
            ws[xlref(count_vlan_row + 1, 1)] = vlan

            for count_col, vlankey in enumerate(vlankeys):
                value = vlanitems.get(vlankey, '')
                if isinstance(value, list):
                    value = ','.join(value)
                ws[xlref(count_vlan_row + 1, count_col + 2)] = value
            count_vlan_row += 1

    if 'OldPortinfo' in sheetnames:
        wb.remove(wb['OldPortinfo'])
    wb.create_sheet('OldPortinfo')
    ws = wb['OldPortinfo']

    count_port_row = 0
    ws[xlref(0, 0)] = 'hostname'
    ws[xlref(0, 1)] = 'interface'
    for count, portkey in enumerate(portkeys):
        ws[xlref(0, count + 2)] = portkey

    for tree in trees:
        for port, portitems in tree['port'].items():
            ws[xlref(count_port_row + 1, 0)] = tree['hostname']
            ws[xlref(count_port_row + 1, 1)] = port

            for count_col, portkey in enumerate(portkeys):
                value = portitems.get(portkey, '')
                if isinstance(value, list):
                    value = ','.join(value)
                ws[xlref(count_port_row + 1, count_col + 2)] = value
            count_port_row += 1

    wb.save(filename)


def short_port_index(port_index):

    temp_index = re.search(r'^[a-zA-z]+(\d+/\d+)', port_index).group(1)
    if 'TenGigabitEthernet' in port_index:
        return 'Te' + temp_index
    elif 'GigabitEthernet' in port_index:
        return 'Gi' + temp_index


def reversed_dict_iter(dict_):

    def interface_id(port):
        return int(port.split('/')[1])

    reversed_dict_keys = sorted(list(dict_), reverse=True, key=interface_id)
    return reversed_dict_keys
