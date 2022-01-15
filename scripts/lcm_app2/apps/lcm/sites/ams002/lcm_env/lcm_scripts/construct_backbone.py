from pol_utils import *
from lcm_scripts import *
from .Stack import Stack, StackMember
import openpyxl

##import sys
##import os
##sys.path.append(os.getcwd() + '\\pol_utils')


def construct_backbone(excel_data, filename):
    """
    This function creates NNI port reservation for the creation of
    EGL-B and EGL-C backbones.

    :param excel_data: dictionary containing data from Excel file.
    :param filename: name of Excel file
    :return: Dictionary with key = device name and value containing stack
             object. This object therefore contains NNI port assignments
             being made in the creation of the backbone. """

    wb = openpyxl.load_workbook(filename)

    site_migration = excel_data['site_migration']
    stackconfig = excel_data['stackconfig']

    hostname_devicerole_map = {}
    for row in site_migration:
        hostname_devicerole_map[row.new_devicename] = row.new_devicerole

    new_devicenames = [row.new_devicename for row in site_migration
                       if row.new_devicename is not None]

    # Construct stacks
    stackconfigs = {}
    for hostname in new_devicenames:
        stack = Stack(hostname_devicerole_map[hostname])
        for row in stackconfig:
            if row.hostname != hostname:
                continue
            # print('test')
            # print(row.devicetype, row.uplinkmodule)
            stackmember = StackMember(row.devicetype, row.uplinkmodule)
            stack.add_stackmember(stackmember)
        stackconfigs[hostname] = stack

    # Print UNI ports per stack in separate tabs
    for hostname in stackconfigs:
        ws = wb.create_sheet(hostname, 0)
        ws[xlref(0, 0)] = 'hostname'
        ws[xlref(0, 1)] = 'interface'
        
        for index, port in enumerate(
                stackconfigs[hostname].get_uni_port_list()):
            ws[xlref(index + 1, 0)] = hostname
            ws[xlref(index + 1, 1)] = port

    # Make NNI port reservations.
    nni_migration = excel_data['nni_migration']
    for row in nni_migration:
        if row.new_hostname is None:
            continue
        stackconfigs[row.new_hostname].reserve_nni_port(row.new_port)

    # Construct backbone
    switches = [hostname for hostname in stackconfigs
                if stackconfigs[hostname].devicerole == 'switch']
    routers = [hostname for hostname in stackconfigs
               if stackconfigs[hostname].devicerole == 'router']

    if len(routers) == 1:  # EGL-B
        for switch in switches:
            stackconfigs[routers[0]].auto_assign_nni_port(switch, order='fwd')
            stackconfigs[routers[0]].auto_assign_nni_port(switch, order='fwd')
            stackconfigs[switch].auto_assign_nni_port(routers[0], order='fwd')
            stackconfigs[switch].auto_assign_nni_port(routers[0], order='fwd')

        for switch in switches:
            routerports = stackconfigs[routers[0]].reserve_ports_for(switch)
            switchports = stackconfigs[switch].reserve_ports_for(routers[0])
            stackconfigs[routers[0]].connect_devices(
                routerports[0], switchports[0]
            )
            stackconfigs[routers[0]].connect_devices(
                routerports[1], switchports[1]
            )
            stackconfigs[switch].connect_devices(
                switchports[0], routerports[0]
            )
            stackconfigs[switch].connect_devices(
                switchports[1], routerports[1]
            )

    elif len(routers) == 2:  # EGL-C
        stackconfigs[routers[0]].auto_assign_nni_port(routers[1])
        stackconfigs[routers[0]].auto_assign_nni_port(routers[1])
        stackconfigs[routers[1]].auto_assign_nni_port(routers[0])
        stackconfigs[routers[1]].auto_assign_nni_port(routers[0])

        routerports0 = stackconfigs[routers[0]].reserve_ports_for(routers[1])
        routerports1 = stackconfigs[routers[1]].reserve_ports_for(routers[0])
        stackconfigs[routers[0]].connect_devices(
            routerports0[0], routerports1[0]
        )
        stackconfigs[routers[0]].connect_devices(
            routerports0[1], routerports1[1]
        )
        stackconfigs[routers[1]].connect_devices(
            routerports0[0], routerports1[0]
        )
        stackconfigs[routers[1]].connect_devices(
            routerports0[1], routerports1[1]
        )

        for router in routers:
            for switch in switches:
                stackconfigs[router].auto_assign_nni_port(switch, order='fwd')
                stackconfigs[switch].auto_assign_nni_port(router, order='fwd')

        for router in routers:
            for switch in switches:
                routerports = stackconfigs[router].reserve_ports_for(switch)
                switchports = stackconfigs[switch].reserve_ports_for(router)
                stackconfigs[router].connect_devices(
                    routerports[0], switchports[0]
                )
                stackconfigs[switch].connect_devices(
                    switchports[0], routerports[0]
                )

    #  Print NNI port info to Excel
    sheetnames = wb.sheetnames
    if 'Backbone' in sheetnames:
        wb.remove(wb['Backbone'])
    wb.create_sheet('Backbone')
    sheet = wb['Backbone']

    sheet[xlref(0, 0)] = 'Hostname'
    sheet[xlref(0, 1)] = 'LocalPort'
    sheet[xlref(0, 2)] = 'RemoteHostname'
    sheet[xlref(0, 3)] = 'RemotePort'

    row = 1
    for router in routers:
        for port, items in stackconfigs[router].nni_port_assignment.items():
            sheet[xlref(row, 0)] = router
            sheet[xlref(row, 1)] = port
            sheet[xlref(row, 2)] = items[0]
            sheet[xlref(row, 3)] = items[1]
            row += 1

    wb.save(filename)
    return stackconfigs
