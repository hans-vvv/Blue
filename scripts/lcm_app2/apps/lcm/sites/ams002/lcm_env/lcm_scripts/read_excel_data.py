from collections import namedtuple
import openpyxl


def read_excel_data(filename):

    """
    This function reads information from several tabs from Excel file.

    :param filename: Excel file
    :return: dictionary containing info from Excel file
    """

    wb = openpyxl.load_workbook(filename)

    # remove all tabs in sheet other then fixed tabs listed here under
    fixed_tab_names = ['SiteMigration', 'OldPortinfo', 'OldVlaninfo',
                       'UniMigrate', 'NniMigrate', 'StackConfig']
    sheetnames = wb.sheetnames
    for sheetname in sheetnames:
        if sheetname not in fixed_tab_names:
            wb.remove(wb[sheetname])

    excel_data = {}  # data from sheets

    # Read VLAN migration data from SiteConfig tab.
    sheet = wb['SiteMigration']

    col = {}
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        col[column[0].value] = index

    SiteMigration = namedtuple(
        'SiteMigration', ['old_devicename', 'old_devicerole',
                          'new_devicename', 'new_devicerole',
                          'new_mgmt_prefix', 'new_mgmt_gw', 'from_vlan',
                          'to_vlan', 'vlan_type', 'base_vlan_name', 'vrf']
    )
    site_migration = []
    for index, row in enumerate(sheet.iter_rows()):
        if index == 0:
            continue
        test = [str(cell.value).strip() for cell in row]
        table_row = [str(cell.value).strip() if cell.value else None
                     for cell in row]
        from_vlan = table_row[col['FromVlan']]
        to_vlan = table_row[col['ToVlan']]
        vlan_type = table_row[col['VlanType']]
        old_devicename = table_row[col['OldDeviceName']]
        old_devicerole = table_row[col['OldDeviceRole']]
        new_devicename = table_row[col['NewDeviceName']]
        new_devicerole = table_row[col['NewDeviceRole']]
        new_mgmt_prefix = table_row[col['NewMgmtPrefix']]
        new_mgmt_gw = table_row[col['NewMgmtGw']]
        base_vlan_name = table_row[col['BaseVlanName']]
        vrf = table_row[col['Vrf']]

        row = SiteMigration(
            old_devicename, old_devicerole, new_devicename, new_devicerole,
            new_mgmt_prefix, new_mgmt_gw, from_vlan, to_vlan, vlan_type,
            base_vlan_name, vrf
        )
        site_migration.append(row)

    excel_data['site_migration'] = site_migration

    # Read stack configurations
    sheet = wb['StackConfig']
    StackHw = namedtuple('StackHw', ['hostname', 'stackmember', 'devicetype',
                                     'uplinkmodule'])
    stackconfig = []

    col = {}
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        col[column[0].value] = index

    for index, row in enumerate(sheet.iter_rows()):
        if index == 0:
            continue
        table_row = [str(cell.value).strip() if cell.value else None
                     for cell in row]

        hostname = table_row[col['SiteConfigNewDeviceName']]
        stackmember = table_row[col['StackMember']]
        devicetype = table_row[col['DeviceType']]
        uplinkmodule = table_row[col['UplinkModule']]

        row = StackHw(hostname, stackmember, devicetype, uplinkmodule)
        stackconfig.append(row)

    excel_data['stackconfig'] = stackconfig

    # Read UNI hostname/port migration table
    sheet = wb['UniMigrate']

    col = {}
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        col[column[0].value] = index
    UniMigration = namedtuple('UniMigration', ['old_hostname', 'old_port',
                                               'new_hostname', 'new_port'])
    uni_migration = []

    for index, row in enumerate(sheet.iter_rows()):
        if index == 0:
            continue
        table_row = [str(cell.value).strip() if cell.value else None
                     for cell in row]

        old_hostname = table_row[col['OldHostname']]
        old_port = table_row[col['OldPort']]
        new_hostname = table_row[col['NewHostname']]
        new_port = table_row[col['NewPort']]

        uni_migration.append(
            UniMigration(old_hostname, old_port, new_hostname, new_port))

    excel_data['uni_migration'] = uni_migration

    # Read NNI migration tab
    sheet = wb['NniMigrate']
    NniMigration = namedtuple('NniMigration', ['old_hostname', 'old_port',
                                               'new_hostname', 'new_port'])
    nni_migration = []

    col = {}
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        col[column[0].value] = index

    for index, row in enumerate(sheet.iter_rows()):
        if index == 0:
            continue
        table_row = [str(cell.value).strip() if cell.value else None
                     for cell in row]

        old_hostname = table_row[col['OldHostname']]
        old_port = table_row[col['OldPort']]
        new_hostname = table_row[col['NewHostname']]
        new_port = table_row[col['NewPort']]

        nni_migration.append(
            NniMigration(old_hostname, old_port, new_hostname, new_port))

    excel_data['nni_migration'] = nni_migration

    wb.save(filename)

    # Read Inventory sheet
    wb = openpyxl.load_workbook('Inventory.xlsx')
    sheet = wb['FlatInventory']
    Inventory = namedtuple('Inventoty', ['hostname', 'ip_address'])

    inventory = []

    col = {}
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        col[column[0].value] = index

    for index, row in enumerate(sheet.iter_rows()):
        if index == 0:
            continue
        table_row = [str(cell.value).strip() if cell.value else None
                     for cell in row]

        hostname = table_row[col['Hostname']]
        ip_address = table_row[col['IP Address']]

        inventory.append(
            Inventory(hostname, ip_address))

    excel_data['inventory'] = inventory
                           
 
    return excel_data


if __name__ == '__main__':

    filename = 'result.xlsx'
    read_excel_data(filename)
