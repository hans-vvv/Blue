from pol_utils import *
import openpyxl


def update_sheet_1(old_config_data, filename):

    """
    This procedure updates the Excel file with data taken from the old network
    elements.

    :param old_config_data: object with configuration data from old devices
    :param filename: Name of Excel file
    :return: None
    """

    # Collect VLAN dB's and write into sheet in SiteMigration tab

    # Store all VLAN's from device VLAN dB's in list.
    vlan_db = set()
    old_hostnames = [tree['hostname'] for tree in old_config_data]
    for old_hostname in old_hostnames:
        tree = get_tree_by_hostname(old_config_data, old_hostname)
        for vlan, items in tree['vlan'].items():
            vlan_db.add(vlan)
    vlan_list = sorted(vlan_db, key=int)
    # print(vlan_list)

    wb = openpyxl.load_workbook(filename)
    sheet = wb['SiteMigration']

    col = {}
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        col[column[0].value] = index

    # Clear data from FromVlan column
    for i in range(1, 30):
        sheet[xlref(i, col['FromVlan'])] = None

    # Insert FromVlan into sheet
    for i, vlan in enumerate(vlan_list):
        sheet[xlref(i + 1, col['FromVlan'])] = int(vlan)

    wb.save(filename)

    # Collect UNI ports from devices with switch role and print
    # in UNI migrate tab
    sheet = wb['UniMigrate']

    wb.save(filename)
